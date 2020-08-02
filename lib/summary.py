import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import json

wb = openpyxl.Workbook()
wb.remove(wb.active)

title_cols = ['filename', 'target', 'time second for encode per frame', 'bitrate', 'vmaf', 'vmaf min', 'vmaf std',
              'psnr', 'psnr min', 'psnr std', 'ssim', 'ssim min', 'ssim std']


def record_task_results(data):
    print(data)
    ws = wb.create_sheet(data['task_name'])
    start_row = 1
    start_col = 1
    # write title
    for row in [start_row]:
        for col in range(start_col, start_col + len(title_cols)):
            _ = ws.cell(column=col, row=row, value=title_cols[col - 1])

    if data['repeat_target']:
        target_col = f'{get_column_letter(2)}${start_row}'
        ws[target_col] = f'{ws[target_col].value}({data["repeat_target"]})'

    for row in ws.iter_rows(min_row=start_row, max_row=start_row, max_col=len(title_cols)):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor='DDDDDD')

    start_row += 1

    for file_item in data['results']:
        file_results = file_item['results']
        for index, result in enumerate(file_results):
            values = [file_item['name'] if index == 0 else '', result['repeat_value'], result['encode_fps'],
                      result['bitrate']]
            for score in result['scores']:
                values = values + score

            for row in [start_row]:
                for col in range(start_col, start_col + len(title_cols)):
                    _ = ws.cell(column=col, row=row, value=values[col - 1])
            start_row += 1
        start_row += 1


def save(path):
    wb.save(path)


if __name__ == '__main__':
    test_json = json.loads(
        '{"task_name":"preset test","repeat_target":"preset","results":[{"name":"desktop_640_360_30.yuv","results":[{'
        '"repeat_value":"fast","bitrate":704.43,"encode_fps":0.00148,"scores":[[100,100,0],[37.71246,34.46934,'
        '0.98309],[0.96504,0.948,0.00483]]},{"repeat_value":"faster","bitrate":674.96,"encode_fps":0.00108,'
        '"scores":[[99.9997,99.8137,0.0063],[37.49444,34.34738,0.95364],[0.96404,0.948,0.00476]]}]}]}')
    record_task_results(test_json)
    save('./test.xlsx')
